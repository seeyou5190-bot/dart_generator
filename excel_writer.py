"""excel_writer.py - 재무제표 엑셀 출력 모듈"""

import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

C_HDR_BG  = "1F4E79"; C_HDR_FG  = "FFFFFF"
C_SUB_BG  = "2E75B6"; C_SUB_FG  = "FFFFFF"
C_SEC_BG  = "D6E4F0"; C_SEC_FG  = "1F4E79"
C_ALT     = "F2F7FC"; C_TOTAL   = "FFF2CC"; C_TOTAL_FG = "7F6000"
C_NEG     = "C00000"; C_BORDER  = "BDD7EE"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size)
def _border():
    s = Side(border_style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="center", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


class ExcelWriter:
    def __init__(self, config):
        self.config      = config
        self.unit_label  = config.get("unit", {}).get("target", "백만원")
        self.unit_div    = {"원":1,"천원":1_000,"백만원":1_000_000,"억원":100_000_000}.get(self.unit_label, 1_000_000)

    def write(self, statements, bs_standards, is_standards, output_path) -> str:
        wb = Workbook()
        wb.remove(wb.active)
        if self.config.get("output", {}).get("include_comparison_sheet", True):
            self._comparison(wb, statements, bs_standards, is_standards)
        for fs in statements:
            self._company_sheet(wb, fs, bs_standards, is_standards)
        errors = [(fs.short_name, e) for fs in statements for e in fs.errors]
        if errors:
            self._error_sheet(wb, errors)
        wb.save(output_path)
        logger.info(f"저장 완료: {output_path}")
        return output_path

    def _comparison(self, wb, statements, bs_stds, is_stds):
        ws  = wb.create_sheet("손보사 비교")
        ws.freeze_panes = "C3"
        cos = [fs for fs in statements if not fs.errors]
        if not cos: return
        period = cos[0].period_label

        # 헤더 행1: 타이틀
        last_col = get_column_letter(2 + len(cos))
        ws.merge_cells(f"A1:{last_col}1")
        c = ws["A1"]
        c.value = f"손해보험사 재무제표 비교  |  {period}  |  (단위: {self.unit_label})"
        c.font = _font(True, C_HDR_FG, 12); c.fill = _fill(C_HDR_BG); c.alignment = _align("left")

        # 헤더 행2: 회사명
        for ci, val in enumerate(["구분", "계정과목"], 1):
            c = ws.cell(2, ci, val)
            c.font = _font(True, C_SUB_FG); c.fill = _fill(C_SUB_BG); c.alignment = _align("center")
        for ci, fs in enumerate(cos, 3):
            cl = ws.cell(2, ci, fs.short_name)
            cl.font = _font(True, C_SUB_FG); cl.fill = _fill(C_SUB_BG)
            cl.alignment = _align(); cl.border = _border()

        # 헤더 행3: 단위 표시
        for ci, fs in enumerate(cos, 3):
            c = ws.cell(3, ci, f"({fs.get_unit_label()})")
            c.font = _font(size=9); c.alignment = _align(); c.border = _border()

        row = 4
        for sec_name, stds, getter in [
            ("재무상태표", bs_stds, lambda fs, a: fs.balance_sheet.get(a)),
            ("손익계산서", is_stds, lambda fs, a: fs.income_stmt.get(a)),
        ]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2+len(cos))
            c = ws.cell(row, 1, sec_name)
            c.font = _font(True, C_SEC_FG, 11); c.fill = _fill(C_SEC_BG)
            c.alignment = _align("left"); c.border = _border()
            row += 1

            for i, acct in enumerate(stds):
                bg     = C_ALT if i % 2 == 0 else "FFFFFF"
                is_tot = "총계" in acct or "합계" in acct
                ws.cell(row, 1).fill   = _fill(bg); ws.cell(row, 1).border = _border()
                ac = ws.cell(row, 2, acct)
                ac.font = _font(is_tot); ac.fill = _fill(C_TOTAL if is_tot else bg)
                ac.alignment = _align("left"); ac.border = _border()
                for ci, fs in enumerate(cos, 3):
                    raw = getter(fs, acct)
                    val = self._conv(raw, fs.unit_won)
                    cell = ws.cell(row, ci)
                    if val is None:
                        cell.value = "-"; cell.alignment = _align()
                    else:
                        cell.value = val; cell.number_format = "#,##0"
                        cell.alignment = _align()
                        if val < 0: cell.font = _font(color=C_NEG)
                    cell.fill = _fill(C_TOTAL if is_tot else bg); cell.border = _border()
                row += 1
            row += 1

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 26
        for ci in range(3, 3 + len(cos)):
            ws.column_dimensions[get_column_letter(ci)].width = 16

    def _company_sheet(self, wb, fs, bs_stds, is_stds):
        ws = wb.create_sheet(fs.short_name[:28])
        ws.freeze_panes = "A3"
        ws.merge_cells("A1:C1")
        c = ws["A1"]
        c.value = f"{fs.company_name}  |  {fs.period_label}  |  (단위: {self.unit_label})"
        c.font = _font(True, C_HDR_FG, 11); c.fill = _fill(C_HDR_BG); c.alignment = _align("left")

        for ci, h in enumerate(["구분", "계정과목", f"금액({self.unit_label})"], 1):
            c = ws.cell(2, ci, h)
            c.font = _font(True, C_SUB_FG); c.fill = _fill(C_SUB_BG)
            c.alignment = _align(); c.border = _border()

        row = 3
        for sec, stds, getter in [
            ("재무상태표", bs_stds, lambda fs, a: fs.balance_sheet.get(a)),
            ("손익계산서", is_stds, lambda fs, a: fs.income_stmt.get(a)),
        ]:
            if not stds: continue
            
            # 섹션 헤더
            ws.merge_cells(f"A{row}:C{row}")
            c = ws.cell(row, 1, sec)
            c.font = _font(True, C_SEC_FG); c.fill = _fill(C_SEC_BG)
            c.alignment = _align("left"); c.border = _border()
            row += 1

            for i, acct in enumerate(stds):
                bg = C_ALT if i % 2 == 0 else "FFFFFF"
                # "합계", "총계" 등이 포함된 행은 강조
                is_tot = any(k in acct for k in ["총계", "합계", "총자산", "총부채", "총자본"])
                
                ws.cell(row, 1).fill = _fill(bg); ws.cell(row, 1).border = _border()
                
                a = ws.cell(row, 2, acct)
                a.font = _font(is_tot); a.fill = _fill(C_TOTAL if is_tot else bg)
                a.alignment = _align("left"); a.border = _border()
                
                val = getter(fs, acct)
                if val is not None:
                    val = self._conv(val, fs.unit_won)
                vc  = ws.cell(row, 3)
                if val is None:
                    vc.value = "-"; vc.alignment = _align()
                else:
                    vc.value = val; vc.number_format = "#,##0"
                    vc.alignment = _align("right")
                    if val < 0: vc.font = _font(color=C_NEG)
                
                vc.fill = _fill(C_TOTAL if is_tot else bg); vc.border = _border()
                row += 1
            row += 1

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 18

    def _error_sheet(self, wb, errors):
        ws = wb.create_sheet("오류 목록")
        ws["A1"] = "회사명"; ws["B1"] = "오류 내용"
        for c in [ws["A1"], ws["B1"]]: c.font = _font(bold=True)
        for i, (co, msg) in enumerate(errors, 2):
            ws.cell(i, 1, co); ws.cell(i, 2, msg)
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 60

    def _conv(self, value, src_unit):
        if value is None: return None
        return (value * src_unit) / self.unit_div
